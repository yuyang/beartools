"""账单归一化服务。"""

from __future__ import annotations

from collections.abc import Callable
from pathlib import Path
import re
from typing import Protocol

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from .agent import analyze_bill_row, resolve_bill_structure, resolve_part_refund_amount
from .models import (
    AnalyzeBillFileResult,
    BillAnalysisResult,
    BillFieldMapping,
    BillNormalizedStatus,
    BillRunProgressState,
    BillStatusMappingConfig,
    BillStructureFileResult,
    NormalizeBillFileResult,
    NormalizeProgressSnapshot,
    NormalizedBillRow,
    PartRefundAmountResult,
    RunBillPipelineResult,
    UnknownBillStatusesError,
)
from .reader import read_bill_preview, read_bill_rows
from .status_mapping import DEFAULT_STATUS_MAPPING_PATH, load_status_mapping, resolve_normalized_status

_AMOUNT_PATTERN = re.compile(r"-?\d+(?:\.\d+)?")
_NORMALIZE_OUTPUT_FIELDNAMES = ["原始来源", "交易时间", "交易对方", "金额", "交易状态", "注意", "备注"]
_ANALYSIS_REQUIRED_HEADERS = ["原始来源", "交易时间", "交易对方", "金额", "交易状态", "注意", "备注"]
_MAX_FAILURE_ROWS = 5


class PartRefundAmountResolver(Protocol):
    """部分退款金额修正器协议。"""

    def __call__(
        self,
        *,
        counterparty: str,
        remark: str,
        status: str,
        amount: str,
        source: str,
        transaction_time: str,
    ) -> PartRefundAmountResult: ...


class NormalizeProgressCallback(Protocol):
    """归一化进度回调协议。"""

    def __call__(self, progress: NormalizeProgressSnapshot) -> None: ...


def normalize_bill_file(
    input_path: str | Path,
    from_value: str,
    *,
    structure_resolver: Callable[[Path], BillStructureFileResult] | None = None,
    part_refund_amount_resolver: PartRefundAmountResolver = resolve_part_refund_amount,
    progress_callback: NormalizeProgressCallback | None = None,
) -> NormalizeBillFileResult:
    """将单个账单文件归一化输出为统一 Excel。"""

    source_path = Path(input_path)
    resolver = structure_resolver or _default_structure_resolver

    # 1. 识别账单结构
    structure = resolver(source_path)
    output_path = _build_output_excel_path(from_value, structure.source)

    # 2. 读取并归一化所有行
    rows = read_bill_rows(source_path)
    unknown_statuses = _collect_unknown_statuses(rows, structure)
    if unknown_statuses:
        raise UnknownBillStatusesError(unknown_statuses)
    normalized_rows, ignored_lines, total_raw_data_rows, _row_numbers = _normalize_rows(
        rows,
        structure,
        from_value=from_value,
        part_refund_amount_resolver=part_refund_amount_resolver,
        progress_callback=progress_callback,
    )

    output_row_count = len(normalized_rows)

    # 3. 写入结果Excel
    _write_normalized_excel(output_path, normalized_rows)

    return NormalizeBillFileResult(
        input_path=source_path,
        output_path=output_path,
        source=structure.source,
        row_count=output_row_count,
        ignored_lines=ignored_lines,
        total_raw_data_rows=total_raw_data_rows,
        output_row_count=output_row_count,
    )


def _default_structure_resolver(input_path: Path) -> BillStructureFileResult:
    preview = read_bill_preview(input_path)
    return resolve_bill_structure(preview)


def _normalize_rows(
    rows: list[list[str]],
    structure: BillStructureFileResult,
    *,
    from_value: str,
    part_refund_amount_resolver: PartRefundAmountResolver,
    progress_callback: NormalizeProgressCallback | None,
) -> tuple[list[NormalizedBillRow], list[int], int, list[int]]:
    if structure.header_row <= 0 or structure.data_start_row <= 0:
        raise RuntimeError("账单结构识别结果缺少有效的 header_row 或 data_start_row")
    if len(rows) < structure.header_row:
        raise RuntimeError("账单文件行数不足，无法定位表头")

    header = rows[structure.header_row - 1]
    column_map = {column_name: index for index, column_name in enumerate(header)}
    field_mapping = structure.field_mapping
    _validate_required_columns(column_map, field_mapping)

    normalized_rows: list[NormalizedBillRow] = []
    row_numbers: list[int] = []  # 记录每个normalized行对应的原始文件行号
    ignored_lines: list[int] = []
    data_rows = rows[structure.data_start_row - 1 :]
    total_raw_data_rows = len(data_rows)
    status_mapping = load_status_mapping(DEFAULT_STATUS_MAPPING_PATH)
    processed_count = 0
    normal_success_count = 0
    refund_count = 0
    part_refund_count = 0
    ignore_count = 0

    for index, row in enumerate(data_rows):
        original_line_number = structure.data_start_row + index
        skip_action = _check_skip_action(
            row=row,
            index=index,
            data_rows=data_rows,
            original_line_number=original_line_number,
            structure=structure,
            column_map=column_map,
            field_mapping=field_mapping,
            ignored_lines=ignored_lines,
        )
        if skip_action == "continue":
            continue
        if skip_action == "break":
            break
        raw_amount = _normalize_amount(_get_column_value(row, column_map, field_mapping.amount.column_name))
        status = _get_column_value(row, column_map, field_mapping.status.column_name)
        resolved_row = _resolve_row_status_and_amount(
            row=row,
            column_map=column_map,
            field_mapping=field_mapping,
            raw_amount=raw_amount,
            status=status,
            source=structure.source,
            status_mapping=status_mapping,
            part_refund_amount_resolver=part_refund_amount_resolver,
        )
        if resolved_row is None:
            ignored_lines.append(original_line_number)
            ignore_count += 1
            continue
        normalized_status, adjusted_amount = resolved_row

        normalized_row = _build_normalized_row(
            row=row,
            column_map=column_map,
            field_mapping=field_mapping,
            from_value=from_value,
            source=structure.source,
            amount=adjusted_amount,
            raw_amount=raw_amount,
            status=status,
            normalized_status=normalized_status,
        )
        normalized_rows.append(normalized_row)
        row_numbers.append(original_line_number)
        processed_count, normal_success_count, refund_count, part_refund_count = _update_normalize_progress_counts(
            processed_count=processed_count,
            normal_success_count=normal_success_count,
            refund_count=refund_count,
            part_refund_count=part_refund_count,
            normalized_status=normalized_status,
        )
        _emit_normalize_progress(
            progress_callback=progress_callback,
            processed_count=processed_count,
            normal_success_count=normal_success_count,
            refund_count=refund_count,
            part_refund_count=part_refund_count,
            ignore_count=ignore_count,
            is_final=False,
        )
    if progress_callback is not None:
        _emit_normalize_progress(
            progress_callback=progress_callback,
            processed_count=processed_count,
            normal_success_count=normal_success_count,
            refund_count=refund_count,
            part_refund_count=part_refund_count,
            ignore_count=ignore_count,
            is_final=True,
        )
    return normalized_rows, ignored_lines, total_raw_data_rows, row_numbers


def _check_skip_action(
    *,
    row: list[str],
    index: int,
    data_rows: list[list[str]],
    original_line_number: int,
    structure: BillStructureFileResult,
    column_map: dict[str, int],
    field_mapping: BillFieldMapping,
    ignored_lines: list[int],
) -> str | None:
    if _is_empty_row(row):
        ignored_lines.append(original_line_number)
        return "continue"
    if not _is_summary_row(row, column_map, field_mapping):
        return None
    ignored_lines.append(original_line_number)
    # 汇总行之后的所有行都忽略
    for remaining_index in range(index + 1, len(data_rows)):
        ignored_lines.append(structure.data_start_row + remaining_index)
    return "break"


def _build_normalized_row(
    *,
    row: list[str],
    column_map: dict[str, int],
    field_mapping: BillFieldMapping,
    from_value: str,
    source: str,
    amount: str,
    raw_amount: str,
    status: str,
    normalized_status: BillNormalizedStatus,
) -> NormalizedBillRow:
    remark = _build_remark(row, column_map, field_mapping.remark_columns.column_names)
    if normalized_status == "PART_REFUND":
        remark = _append_original_amount_remark(remark, raw_amount)
    return NormalizedBillRow(
        from_value=from_value,
        source=source,
        transaction_time=_get_column_value(row, column_map, field_mapping.transaction_time.column_name),
        counterparty=_get_column_value(row, column_map, field_mapping.counterparty.column_name),
        amount=amount,
        status=status,
        normalized_status=normalized_status,
        remark=remark,
    )


def _append_original_amount_remark(remark: str, raw_amount: str) -> str:
    if not raw_amount:
        return remark
    original_amount_part = f"原始金额={raw_amount}"
    if not remark:
        return original_amount_part
    return f"{remark}; {original_amount_part}"


def _is_zero_amount(amount: str) -> bool:
    return float(amount) == 0.0


def _should_ignore_part_refund(normalized_status: BillNormalizedStatus, adjusted_amount: str) -> bool:
    return normalized_status == "PART_REFUND" and _is_zero_amount(adjusted_amount)


def _should_ignore_status(normalized_status: BillNormalizedStatus) -> bool:
    return normalized_status == "IGNORE"


def _resolve_row_status_and_amount(
    *,
    row: list[str],
    column_map: dict[str, int],
    field_mapping: BillFieldMapping,
    raw_amount: str,
    status: str,
    source: str,
    status_mapping: BillStatusMappingConfig,
    part_refund_amount_resolver: PartRefundAmountResolver,
) -> tuple[BillNormalizedStatus, str] | None:
    normalized_status = resolve_normalized_status(status, status_mapping)
    if normalized_status is None:
        raise RuntimeError(f"未识别的交易状态: {status}")
    if _should_ignore_status(normalized_status):
        return None

    adjusted_amount = raw_amount
    try:
        adjusted_amount = _normalize_row_amount(
            normalized_status=normalized_status,
            raw_amount=raw_amount,
            row=row,
            column_map=column_map,
            field_mapping=field_mapping,
            status=status,
            source=source,
            transaction_time=_get_column_value(row, column_map, field_mapping.transaction_time.column_name),
            part_refund_amount_resolver=part_refund_amount_resolver,
        )
    except (ValueError, TypeError):
        # 不是合法数字，保持原始值
        pass
    if _should_ignore_part_refund(normalized_status, adjusted_amount):
        return None
    return normalized_status, adjusted_amount


def _update_normalize_progress_counts(
    *,
    processed_count: int,
    normal_success_count: int,
    refund_count: int,
    part_refund_count: int,
    normalized_status: BillNormalizedStatus,
) -> tuple[int, int, int, int]:
    processed_count += 1
    if normalized_status == "NORMAL_SUCCESS":
        normal_success_count += 1
    elif normalized_status == "REFUND":
        refund_count += 1
    else:
        part_refund_count += 1
    return processed_count, normal_success_count, refund_count, part_refund_count


def _emit_normalize_progress(
    *,
    progress_callback: NormalizeProgressCallback | None,
    processed_count: int,
    normal_success_count: int,
    refund_count: int,
    part_refund_count: int,
    ignore_count: int,
    is_final: bool,
) -> None:
    if progress_callback is None:
        return
    if not is_final and processed_count % 100 != 0:
        return
    progress_callback(
        NormalizeProgressSnapshot(
            processed_count=processed_count,
            normal_success_count=normal_success_count,
            refund_count=refund_count,
            part_refund_count=part_refund_count,
            ignore_count=ignore_count,
            is_final=is_final,
        )
    )


def _is_empty_row(row: list[str]) -> bool:
    return not any(cell.strip() for cell in row)


def _get_column_value(row: list[str], column_map: dict[str, int], column_name: str) -> str:
    if not column_name:
        return ""
    index = column_map.get(column_name)
    if index is None or index >= len(row):
        return ""
    return row[index].strip()


def _normalize_amount(raw_amount: str) -> str:
    if not raw_amount:
        return ""
    match = _AMOUNT_PATTERN.search(raw_amount.replace(",", ""))
    if match is None:
        return raw_amount.strip()
    return match.group(0)


def _build_remark(row: list[str], column_map: dict[str, int], column_names: list[str]) -> str:
    parts: list[str] = []
    for column_name in column_names:
        value = _get_column_value(row, column_map, column_name)
        if not value:
            continue
        parts.append(f"{column_name}={value}")
    return "; ".join(parts)


def _build_output_excel_path(from_value: str, source: str) -> Path:
    return Path("data") / "bill" / f"{from_value}{source}.xlsx"


def _collect_unknown_statuses(rows: list[list[str]], structure: BillStructureFileResult) -> list[str]:
    header = rows[structure.header_row - 1]
    column_map = {column_name: index for index, column_name in enumerate(header)}
    status_column = structure.field_mapping.status.column_name
    mapping = load_status_mapping(DEFAULT_STATUS_MAPPING_PATH)
    unknown_statuses: list[str] = []
    seen: set[str] = set()
    for row in rows[structure.data_start_row - 1 :]:
        if _is_empty_row(row):
            continue
        raw_status = _get_column_value(row, column_map, status_column)
        if not raw_status:
            continue
        if resolve_normalized_status(raw_status, mapping) is None and raw_status not in seen:
            seen.add(raw_status)
            unknown_statuses.append(raw_status)
    return unknown_statuses


def _normalize_row_amount(
    *,
    normalized_status: BillNormalizedStatus,
    raw_amount: str,
    row: list[str],
    column_map: dict[str, int],
    field_mapping: BillFieldMapping,
    status: str,
    source: str,
    transaction_time: str,
    part_refund_amount_resolver: PartRefundAmountResolver,
) -> str:
    amount_num = float(raw_amount)
    if normalized_status == "NORMAL_SUCCESS":
        return f"{abs(amount_num):g}"
    if normalized_status == "REFUND":
        return f"{-abs(amount_num):g}"

    part_refund_result = part_refund_amount_resolver(
        counterparty=_get_column_value(row, column_map, field_mapping.counterparty.column_name),
        remark=_build_remark(row, column_map, field_mapping.remark_columns.column_names),
        status=status,
        amount=raw_amount,
        source=source,
        transaction_time=transaction_time,
    )
    refund_amount = abs(float(part_refund_result.refund_amount))
    net_amount = abs(amount_num) - refund_amount
    return f"{net_amount:g}"


def _write_normalized_excel(output_path: Path, rows: list[NormalizedBillRow]) -> None:
    """将归一化后的行写入Excel文件"""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Excel工作表创建失败")
    ws.title = "账单明细"

    # 写入中文表头
    ws.append(_NORMALIZE_OUTPUT_FIELDNAMES)

    # 写入行数据
    for row in rows:
        notice = "" if row.normalized_status in ["NORMAL_SUCCESS", "REFUND", "PART_REFUND"] else "focus"
        ws.append(
            [  # type: ignore[misc]
                f"{row.from_value}{row.source}",
                row.transaction_time,
                row.counterparty,
                row.amount,
                row.status,
                notice,
                row.remark,
            ]
        )

    # 保存文件
    wb.save(output_path)


def _validate_required_columns(column_map: dict[str, int], field_mapping: BillFieldMapping) -> None:
    required_columns = {
        "transaction_time": field_mapping.transaction_time.column_name,
        "counterparty": field_mapping.counterparty.column_name,
        "amount": field_mapping.amount.column_name,
        "status": field_mapping.status.column_name,
    }
    for field_name, column_name in required_columns.items():
        if not column_name:
            raise RuntimeError(f"账单结构识别结果缺少字段映射: {field_name}")
        if column_name not in column_map:
            raise RuntimeError(f"账单结构识别结果中的列不存在: {column_name}")


def _is_summary_row(row: list[str], column_map: dict[str, int], field_mapping: BillFieldMapping) -> bool:
    transaction_time = _get_column_value(row, column_map, field_mapping.transaction_time.column_name)
    counterparty = _get_column_value(row, column_map, field_mapping.counterparty.column_name)
    amount = _get_column_value(row, column_map, field_mapping.amount.column_name)
    status = _get_column_value(row, column_map, field_mapping.status.column_name)
    if amount and _AMOUNT_PATTERN.search(amount.replace(",", "")) is None:
        return True
    if transaction_time and not _looks_like_transaction_time(transaction_time):
        return True
    return not any([counterparty, status]) and bool(transaction_time) and not amount


def _looks_like_transaction_time(value: str) -> bool:
    normalized = value.strip()
    return bool(re.fullmatch(r"\d{4}-\d{2}-\d{2}(?: \d{2}:\d{2}(?::\d{2})?)?", normalized))


def analyze_bill_file(
    input_path: str | Path,
    *,
    row_analyzer: Callable[[str, str, str, str], BillAnalysisResult] = analyze_bill_row,
    progress_state: BillRunProgressState | None = None,
) -> AnalyzeBillFileResult:
    """分析归一化后的账单文件，每行分析得到用途和归属人，追加列输出到新文件。

    Args:
        input_path: 归一化结果的xlsx文件路径
        row_analyzer: 自定义行分析函数，接收(交易对方, 备注, 交易状态, 金额)，返回分析结果，
            默认使用analyze_bill_row进行LLM分析
        progress_state: 进度状态对象，用于跟踪处理进度

    Returns:
        分析结果对象

    Raises:
        ValueError: 如果输入不是xlsx
        RuntimeError: 如果表头不匹配，或失败行数超过阈值
    """
    input_path = Path(input_path)

    # 更新进度状态
    if progress_state is not None:
        progress_state.current_step = "Analysis"

    # 1. 验证输入格式
    if input_path.suffix.lower() != ".xlsx":
        raise ValueError("只支持归一化结果 .xlsx 输入")

    # 2. 加载工作簿并验证表头
    wb = load_workbook(input_path)
    ws = wb.active
    if ws is None or ws.max_row < 1:
        raise RuntimeError("输入文件为空工作表")

    headers = _load_and_validate_headers(ws)
    column_index_map = {header: idx for idx, header in enumerate(headers)}
    processed_rows = _prepare_output_headers(headers)

    if progress_state is not None:
        progress_state.analysis_total_count = max(ws.max_row - 1, 0)

    # 3. 处理所有数据行
    total_rows, failed_count = _process_data_rows(
        ws, column_index_map, processed_rows, row_analyzer, input_path, progress_state
    )

    # 4. 写入输出文件
    output_path = _write_analysis_output(input_path, processed_rows)

    return AnalyzeBillFileResult(
        input_path=input_path,
        output_path=output_path,
        total_rows=total_rows,
        failed_rows=failed_count,
    )


def _load_and_validate_headers(ws: Worksheet) -> list[str]:
    """加载并验证分析所需的必填表头。"""
    headers: list[str] = []
    for cell in ws[1]:
        if cell.value is not None:
            headers.append(str(cell.value).strip())
        else:
            headers.append("")

    missing_headers = [h for h in _ANALYSIS_REQUIRED_HEADERS if h not in headers]
    if missing_headers:
        raise RuntimeError(
            f"输入文件表头不匹配，需要包含固定表头: {', '.join(_ANALYSIS_REQUIRED_HEADERS)}，缺少: {', '.join(missing_headers)}"
        )
    return headers


def _prepare_output_headers(headers: list[str]) -> list[list[str | None]]:
    """准备输出表头，追加 purpose 和 owner 列。"""
    output_headers: list[str | None] = list(headers)
    if "purpose" not in output_headers:
        output_headers.append("purpose")
    if "owner" not in output_headers:
        output_headers.append("owner")
    return [output_headers]


def _process_data_rows(
    ws: Worksheet,
    column_index_map: dict[str, int],
    processed_rows: list[list[str | None]],
    row_analyzer: Callable[[str, str, str, str], BillAnalysisResult],
    input_path: Path,
    progress_state: BillRunProgressState | None = None,
) -> tuple[int, int]:
    """处理所有数据行，执行分析并收集结果。"""
    failed_count = 0
    total_rows = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        total_rows += 1
        original_values = _extract_original_row_values(row)

        # 提取需要分析的字段
        counterparty = original_values[column_index_map["交易对方"]] or ""
        remark = original_values[column_index_map["备注"]] or ""
        status = original_values[column_index_map["交易状态"]] or ""
        amount = original_values[column_index_map["金额"]] or ""

        # 分析当前行
        purpose, owner = _analyze_single_row(counterparty, remark, status, amount, row_analyzer)

        if purpose == "unknow" and owner == "unknow":
            failed_count += 1

        # 检查失败阈值
        if failed_count > _MAX_FAILURE_ROWS:
            output_path = input_path.with_name(f"{input_path.stem}.analysis.xlsx")
            if output_path.exists():
                output_path.unlink()
            raise RuntimeError(f"分析失败行数超过阈值({_MAX_FAILURE_ROWS})，已达到 {failed_count} 行，终止处理")

        # 追加分析结果
        original_values.append(purpose)
        original_values.append(owner)
        processed_rows.append(original_values)

        # 更新已完成行数计数
        if progress_state is not None:
            progress_state.analysis_completed_count += 1

    return total_rows, failed_count


def _extract_original_row_values(row: tuple[Cell | MergedCell, ...]) -> list[str | None]:
    """提取原始行的单元格值，转换为字符串并去除首尾空白。"""
    original_values: list[str | None] = []
    for cell in row:
        if cell.value is None:
            original_values.append(None)
        else:
            original_values.append(str(cell.value).strip())
    return original_values


def _analyze_single_row(
    counterparty: str,
    remark: str,
    status: str,
    amount: str,
    row_analyzer: Callable[[str, str, str, str], BillAnalysisResult],
) -> tuple[str, str]:
    """分析单行账单数据，返回(用途, 归属人)。"""
    purpose: str
    owner: str
    try:
        analysis = row_analyzer(counterparty, remark, status, amount)
        purpose = analysis.purpose
        owner = analysis.owner
    except Exception:
        purpose = "unknow"
        owner = "unknow"
    return purpose, owner


def _write_analysis_output(input_path: Path, processed_rows: list[list[str | None]]) -> Path:
    """将分析结果写入输出Excel文件。"""
    output_path = input_path.with_name(f"{input_path.stem}.analysis.xlsx")
    output_wb = Workbook()
    output_ws = output_wb.active
    if output_ws is None:
        raise RuntimeError("创建输出Excel工作表失败")

    for row_data in processed_rows:
        output_ws.append(row_data)

    output_wb.save(output_path)
    return output_path


def run_bill_pipeline(
    input_path: str | Path,
    from_value: str,
    *,
    structure_resolver: Callable[[Path], BillStructureFileResult] | None = None,
    row_analyzer: Callable[[str, str, str, str], BillAnalysisResult] = analyze_bill_row,
    progress_state: BillRunProgressState | None = None,
    normalize_progress_callback: NormalizeProgressCallback | None = None,
) -> RunBillPipelineResult:
    """串联 normalize 与 analysis，完成从原始账单到最终分析的完整流程。"""
    input_path_obj = Path(input_path)

    # 初始化进度状态
    if progress_state is not None:
        progress_state.current_step = "Normalize"
        progress_state.analysis_total_count = 0
        progress_state.analysis_completed_count = 0

    # 第一步 normalize
    normalize_result = normalize_bill_file(
        input_path_obj,
        from_value,
        structure_resolver=structure_resolver,
        progress_callback=normalize_progress_callback,
    )

    # 第二步 analysis，确保失败时清理分析输出
    try:
        analyze_result = analyze_bill_file(
            normalize_result.output_path, row_analyzer=row_analyzer, progress_state=progress_state
        )
    except Exception:
        candidate = normalize_result.output_path.with_suffix("").with_suffix(".analysis.xlsx")
        if candidate.exists():
            candidate.unlink()
        raise

    # 标记完成
    if progress_state is not None:
        progress_state.current_step = "Finished"

    return RunBillPipelineResult(
        input_path=input_path_obj,
        normalized_output_path=normalize_result.output_path,
        analysis_output_path=analyze_result.output_path,
        source=normalize_result.source,
        normalized_row_count=normalize_result.output_row_count,
        analysis_total_rows=analyze_result.total_rows,
        analysis_failed_rows=analyze_result.failed_rows,
    )
