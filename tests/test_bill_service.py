from __future__ import annotations

import os
from pathlib import Path

from openpyxl import Workbook


def test_normalize_bill_file_with_csv(tmp_path: Path) -> None:
    from beartools.bill.models import BillFieldDetail, BillFieldMapping, BillRemarkColumns, BillStructureFileResult
    from beartools.bill.service import normalize_bill_file

    input_path = tmp_path / "alipay.csv"
    input_path.write_text(
        "导出信息\n交易时间,交易分类,交易对方,金额,交易状态,收/支,备注\n"
        "2025-12-31 18:54:05,日用百货,山姆会员商店,289.80,交易成功,支出,测试备注\n",
        encoding="utf-8",
    )

    structure = BillStructureFileResult(
        file_name=input_path.name,
        source="支付宝",
        header_row=2,
        data_start_row=3,
        field_mapping=BillFieldMapping(
            transaction_time=BillFieldDetail(column_name="交易时间", confidence="high", reason=""),
            counterparty=BillFieldDetail(column_name="交易对方", confidence="high", reason=""),
            amount=BillFieldDetail(column_name="金额", confidence="high", reason=""),
            status=BillFieldDetail(column_name="交易状态", confidence="high", reason=""),
            remark_columns=BillRemarkColumns(
                column_names=["交易分类", "收/支", "备注"],
                confidence="high",
                reason="",
            ),
        ),
        sample_rows=[],
        notes=[],
    )

    cwd = Path.cwd()
    os.chdir(tmp_path)
    try:
        result = normalize_bill_file(input_path, "2601-", structure_resolver=lambda _: structure)

        assert result.source == "支付宝"
        assert result.row_count == 1
        assert result.output_path == Path("data/bill/2601-支付宝.xlsx")
    finally:
        os.chdir(cwd)


def test_normalize_bill_file_with_xlsx(tmp_path: Path) -> None:
    from beartools.bill.models import BillFieldDetail, BillFieldMapping, BillRemarkColumns, BillStructureFileResult
    from beartools.bill.service import normalize_bill_file

    input_path = tmp_path / "wechat.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "微信支付账单明细"
    sheet.append(["交易时间", "交易类型", "交易对方", "金额(元)", "当前状态", "支付方式", "备注"])
    sheet.append(["2026-01-01 19:43:53", "商户消费", "京东七鲜超市", "¥61.60", "支付成功", "零钱", "/"])
    workbook.save(input_path)

    structure = BillStructureFileResult(
        file_name=input_path.name,
        source="微信",
        header_row=2,
        data_start_row=3,
        field_mapping=BillFieldMapping(
            transaction_time=BillFieldDetail(column_name="交易时间", confidence="high", reason=""),
            counterparty=BillFieldDetail(column_name="交易对方", confidence="high", reason=""),
            amount=BillFieldDetail(column_name="金额(元)", confidence="high", reason=""),
            status=BillFieldDetail(column_name="当前状态", confidence="high", reason=""),
            remark_columns=BillRemarkColumns(
                column_names=["交易类型", "支付方式", "备注"],
                confidence="high",
                reason="",
            ),
        ),
        sample_rows=[],
        notes=[],
    )

    cwd = Path.cwd()
    os.chdir(tmp_path)
    try:
        result = normalize_bill_file(input_path, "2601-", structure_resolver=lambda _: structure)

        assert result.source == "微信"
        assert result.output_path == Path("data/bill/2601-微信.xlsx")
    finally:
        os.chdir(cwd)


def test_normalize_bill_file_maps_status_and_preserves_raw_status(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    from beartools.bill.models import BillFieldDetail, BillFieldMapping, BillRemarkColumns, BillStructureFileResult
    from beartools.bill.service import normalize_bill_file

    input_path = tmp_path / "wechat.csv"
    input_path.write_text(
        "标题\n交易时间,交易对方,金额,当前状态,备注\n2026-01-01 19:43:53,京东,61.60,支付成功,测试\n",
        encoding="utf-8",
    )
    structure = BillStructureFileResult(
        file_name=input_path.name,
        source="微信",
        header_row=2,
        data_start_row=3,
        field_mapping=BillFieldMapping(
            transaction_time=BillFieldDetail(column_name="交易时间", confidence="high", reason=""),
            counterparty=BillFieldDetail(column_name="交易对方", confidence="high", reason=""),
            amount=BillFieldDetail(column_name="金额", confidence="high", reason=""),
            status=BillFieldDetail(column_name="当前状态", confidence="high", reason=""),
            remark_columns=BillRemarkColumns(column_names=["备注"], confidence="high", reason=""),
        ),
        sample_rows=[],
        notes=[],
    )

    cwd = Path.cwd()
    os.chdir(tmp_path)
    try:
        result = normalize_bill_file(input_path, "2601-", structure_resolver=lambda _: structure)
    finally:
        os.chdir(cwd)

    wb = load_workbook(tmp_path / result.output_path)
    ws = wb.active
    assert ws["D2"].value == "61.6"
    assert ws["E2"].value == "支付成功"


def test_normalize_bill_file_raises_unknown_statuses_before_writing_output(tmp_path: Path) -> None:
    import pytest

    from beartools.bill.models import BillFieldDetail, BillFieldMapping, BillRemarkColumns, BillStructureFileResult
    from beartools.bill.service import normalize_bill_file

    input_path = tmp_path / "alipay.csv"
    input_path.write_text(
        "标题\n交易时间,交易对方,金额,交易状态,备注\n2026-01-01 10:00:00,淘宝,10.00,等待确认收货,备注\n",
        encoding="utf-8",
    )
    structure = BillStructureFileResult(
        file_name=input_path.name,
        source="支付宝",
        header_row=2,
        data_start_row=3,
        field_mapping=BillFieldMapping(
            transaction_time=BillFieldDetail(column_name="交易时间", confidence="high", reason=""),
            counterparty=BillFieldDetail(column_name="交易对方", confidence="high", reason=""),
            amount=BillFieldDetail(column_name="金额", confidence="high", reason=""),
            status=BillFieldDetail(column_name="交易状态", confidence="high", reason=""),
            remark_columns=BillRemarkColumns(column_names=["备注"], confidence="high", reason=""),
        ),
        sample_rows=[],
        notes=[],
    )

    cwd = Path.cwd()
    os.chdir(tmp_path)
    try:
        with pytest.raises(RuntimeError, match="等待确认收货"):
            normalize_bill_file(input_path, "2601-", structure_resolver=lambda _: structure)
    finally:
        os.chdir(cwd)


def test_normalize_bill_file_uses_part_refund_amount_resolver(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    from beartools.bill.models import (
        BillFieldDetail,
        BillFieldMapping,
        BillRemarkColumns,
        BillStructureFileResult,
        PartRefundAmountResult,
    )
    from beartools.bill.service import normalize_bill_file

    input_path = tmp_path / "wechat.csv"
    input_path.write_text(
        "标题\n交易时间,交易对方,金额,当前状态,备注\n2026-01-01 19:43:53,京东,61.60,已退款￥1.00,状态显示已退款￥1.00\n",
        encoding="utf-8",
    )
    structure = BillStructureFileResult(
        file_name=input_path.name,
        source="微信",
        header_row=2,
        data_start_row=3,
        field_mapping=BillFieldMapping(
            transaction_time=BillFieldDetail(column_name="交易时间", confidence="high", reason=""),
            counterparty=BillFieldDetail(column_name="交易对方", confidence="high", reason=""),
            amount=BillFieldDetail(column_name="金额", confidence="high", reason=""),
            status=BillFieldDetail(column_name="当前状态", confidence="high", reason=""),
            remark_columns=BillRemarkColumns(column_names=["备注"], confidence="high", reason=""),
        ),
        sample_rows=[],
        notes=[],
    )

    cwd = Path.cwd()
    os.chdir(tmp_path)
    try:
        result = normalize_bill_file(
            input_path,
            "2601-",
            structure_resolver=lambda _: structure,
            part_refund_amount_resolver=lambda **_: PartRefundAmountResult(refund_amount="1.00", reason="命中状态"),
        )
    finally:
        os.chdir(cwd)

    wb = load_workbook(tmp_path / result.output_path)
    ws = wb.active
    assert ws["D2"].value == "60.6"
    assert ws["E2"].value == "已退款￥1.00"
    assert ws["G2"].value == "备注=状态显示已退款￥1.00; 原始金额=61.60"


def test_normalize_bill_file_reports_progress_every_100_rows_and_final(tmp_path: Path) -> None:
    from beartools.bill.models import BillFieldDetail, BillFieldMapping, BillRemarkColumns, BillStructureFileResult
    from beartools.bill.service import normalize_bill_file

    input_path = tmp_path / "wechat.csv"
    header = "标题\n交易时间,交易对方,金额,当前状态,备注\n"
    rows: list[str] = []
    for index in range(105):
        status = "支付成功" if index < 100 else "退款成功"
        rows.append(f"2026-01-01 19:43:{index % 60:02d},京东,1.00,{status},测试")
    input_path.write_text(header + "\n".join(rows) + "\n", encoding="utf-8")

    structure = BillStructureFileResult(
        file_name=input_path.name,
        source="微信",
        header_row=2,
        data_start_row=3,
        field_mapping=BillFieldMapping(
            transaction_time=BillFieldDetail(column_name="交易时间", confidence="high", reason=""),
            counterparty=BillFieldDetail(column_name="交易对方", confidence="high", reason=""),
            amount=BillFieldDetail(column_name="金额", confidence="high", reason=""),
            status=BillFieldDetail(column_name="当前状态", confidence="high", reason=""),
            remark_columns=BillRemarkColumns(column_names=["备注"], confidence="high", reason=""),
        ),
        sample_rows=[],
        notes=[],
    )
    snapshots: list[tuple[int, int, int, int, bool]] = []

    cwd = Path.cwd()
    os.chdir(tmp_path)
    try:
        normalize_bill_file(
            input_path,
            "2601-",
            structure_resolver=lambda _: structure,
            progress_callback=lambda progress: snapshots.append(
                (
                    progress.processed_count,
                    progress.normal_success_count,
                    progress.refund_count,
                    progress.part_refund_count,
                    progress.is_final,
                )
            ),
        )
    finally:
        os.chdir(cwd)

    assert snapshots[0] == (100, 100, 0, 0, False)
    assert snapshots[-1] == (105, 100, 5, 0, True)


def test_normalize_bill_file_ignores_rows_mapped_to_ignore_status(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    from beartools.bill.models import BillFieldDetail, BillFieldMapping, BillRemarkColumns, BillStructureFileResult
    from beartools.bill.service import normalize_bill_file

    input_path = tmp_path / "wechat.csv"
    input_path.write_text(
        "标题\n交易时间,交易对方,金额,当前状态,备注\n"
        "2026-01-01 10:00:00,零钱,5.00,已存入零钱,转入零钱\n"
        "2026-01-01 10:05:00,商家A,10.00,支付成功,正常保留\n"
        "2026-01-01 10:10:00,商家B,20.00,已全额退款,整单退款\n",
        encoding="utf-8",
    )
    structure = BillStructureFileResult(
        file_name=input_path.name,
        source="微信",
        header_row=2,
        data_start_row=3,
        field_mapping=BillFieldMapping(
            transaction_time=BillFieldDetail(column_name="交易时间", confidence="high", reason=""),
            counterparty=BillFieldDetail(column_name="交易对方", confidence="high", reason=""),
            amount=BillFieldDetail(column_name="金额", confidence="high", reason=""),
            status=BillFieldDetail(column_name="当前状态", confidence="high", reason=""),
            remark_columns=BillRemarkColumns(column_names=["备注"], confidence="high", reason=""),
        ),
        sample_rows=[],
        notes=[],
    )

    cwd = Path.cwd()
    os.chdir(tmp_path)
    try:
        result = normalize_bill_file(input_path, "2601-", structure_resolver=lambda _: structure)
    finally:
        os.chdir(cwd)

    assert result.row_count == 1
    assert result.ignored_lines == [3, 5]

    wb = load_workbook(tmp_path / result.output_path)
    ws = wb.active
    assert ws.max_row == 2
    assert ws["C2"].value == "商家A"
    assert ws["E2"].value == "支付成功"


def test_preview_reader_handles_gb18030_csv(tmp_path: Path) -> None:
    from beartools.bill.reader import read_bill_preview

    input_path = tmp_path / "alipay_gbk.csv"
    content = "导出信息\n交易时间,交易对方,金额,交易状态\n2025-12-31 18:54:05,山姆会员商店,289.80,交易成功\n"
    input_path.write_bytes(content.encode("gb18030"))

    preview = read_bill_preview(input_path, max_rows=3)

    assert preview.file_name == input_path.name
    assert preview.file_type == "csv"
    assert "交易时间,交易对方,金额,交易状态" in preview.file_content


def test_preview_reader_escapes_commas_and_newlines(tmp_path: Path) -> None:
    from beartools.bill.reader import read_bill_preview

    input_path = tmp_path / "complex.csv"
    input_path.write_text(
        '交易时间,交易对方,商品说明\n2025-12-31 18:54:05,山姆会员商店,"商品A,包含逗号\n第二行"\n',
        encoding="utf-8",
    )

    preview = read_bill_preview(input_path, max_rows=3)

    assert "1: 交易时间,交易对方,商品说明" in preview.file_content
    assert '2: 2025-12-31 18:54:05,山姆会员商店,"商品A,包含逗号\\n第二行"' in preview.file_content


def test_normalize_bill_file_skips_footer_summary_rows(tmp_path: Path) -> None:
    from beartools.bill.models import BillFieldDetail, BillFieldMapping, BillRemarkColumns, BillStructureFileResult
    from beartools.bill.service import normalize_bill_file

    input_path = tmp_path / "alipay.csv"
    input_path.write_text(
        "导出信息\n交易时间,交易分类,交易对方,金额,交易状态,收/支,备注\n"
        "2025-12-31 18:54:05,日用百货,山姆会员商店,289.80,交易成功,支出,测试备注\n"
        "共1笔记录, , , , , , \n",
        encoding="utf-8",
    )

    structure = BillStructureFileResult(
        file_name=input_path.name,
        source="支付宝",
        header_row=2,
        data_start_row=3,
        field_mapping=BillFieldMapping(
            transaction_time=BillFieldDetail(column_name="交易时间", confidence="high", reason=""),
            counterparty=BillFieldDetail(column_name="交易对方", confidence="high", reason=""),
            amount=BillFieldDetail(column_name="金额", confidence="high", reason=""),
            status=BillFieldDetail(column_name="交易状态", confidence="high", reason=""),
            remark_columns=BillRemarkColumns(
                column_names=["交易分类", "收/支", "备注"],
                confidence="high",
                reason="",
            ),
        ),
        sample_rows=[],
        notes=[],
    )

    cwd = Path.cwd()
    os.chdir(tmp_path)
    try:
        result = normalize_bill_file(input_path, "2601-", structure_resolver=lambda _: structure)
    finally:
        os.chdir(cwd)

    assert result.row_count == 1


def test_apply_refund_offset_ignores_full_refund_only() -> None:
    from beartools.bill.models import NormalizedBillRow
    from beartools.bill.service import _apply_refund_offset

    rows = [
        NormalizedBillRow("2601-", "支付宝", "2026-01-01", "淘宝", "10", "交易成功", "NORMAL_SUCCESS", ""),
        NormalizedBillRow("2601-", "支付宝", "2026-01-02", "淘宝", "-10", "退款成功", "REFUND", ""),
        NormalizedBillRow("2601-", "微信", "2026-01-03", "京东", "-1", "已退款￥1.00", "PART_REFUND", ""),
    ]

    filtered_rows, ignored = _apply_refund_offset(rows, [3, 4, 5], [])

    assert len(filtered_rows) == 1
    assert filtered_rows[0].normalized_status == "PART_REFUND"
    assert ignored == [3, 4]


def test_normalize_bill_file_raises_when_mapped_column_missing(tmp_path: Path) -> None:
    from beartools.bill.models import BillFieldDetail, BillFieldMapping, BillRemarkColumns, BillStructureFileResult
    from beartools.bill.service import normalize_bill_file

    input_path = tmp_path / "alipay.csv"
    input_path.write_text(
        "导出信息\n交易时间,交易分类,交易对方,金额,交易状态,收/支,备注\n"
        "2025-12-31 18:54:05,日用百货,山姆会员商店,289.80,交易成功,支出,测试备注\n",
        encoding="utf-8",
    )

    structure = BillStructureFileResult(
        file_name=input_path.name,
        source="支付宝",
        header_row=2,
        data_start_row=3,
        field_mapping=BillFieldMapping(
            transaction_time=BillFieldDetail(column_name="交易时间", confidence="high", reason=""),
            counterparty=BillFieldDetail(column_name="不存在的交易对方列", confidence="high", reason=""),
            amount=BillFieldDetail(column_name="金额", confidence="high", reason=""),
            status=BillFieldDetail(column_name="交易状态", confidence="high", reason=""),
            remark_columns=BillRemarkColumns(column_names=["交易分类"], confidence="high", reason=""),
        ),
        sample_rows=[],
        notes=[],
    )

    import pytest

    cwd = Path.cwd()
    os.chdir(tmp_path)
    try:
        with pytest.raises(RuntimeError, match="不存在的交易对方列"):
            normalize_bill_file(input_path, "2601-", structure_resolver=lambda _: structure)
    finally:
        os.chdir(cwd)


def test_analyze_bill_file_success_adds_purpose_owner_columns(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    from beartools.bill.models import BillAnalysisResult
    from beartools.bill.service import analyze_bill_file

    # 创建一个符合normalize格式的测试xlsx
    input_path = tmp_path / "test.xlsx"
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["原始来源", "交易时间", "交易对方", "金额", "交易状态", "注意", "备注"])
    ws.append(["2601-支付宝", "2025-12-31 18:54:05", "山姆会员商店", "289.80", "交易成功", "", "日用百货"])
    wb.save(input_path)

    def mock_row_analyzer(counterparty: str, remark: str, status: str, amount: str) -> BillAnalysisResult:
        return BillAnalysisResult(purpose="购物消费/日用百货", owner="vv")

    result = analyze_bill_file(input_path, row_analyzer=mock_row_analyzer)
    assert result.total_rows == 1
    assert result.failed_rows == 0
    assert result.output_path.exists()

    # 验证输出包含purpose和owner列
    wb_out = load_workbook(result.output_path)
    ws_out = wb_out.active
    headers = [cell.value for cell in ws_out[1]]
    assert headers == ["原始来源", "交易时间", "交易对方", "金额", "交易状态", "注意", "备注", "purpose", "owner"]
    row2 = [cell.value for cell in ws_out[2]]
    assert row2[-2] == "购物消费/日用百货"
    assert row2[-1] == "vv"


def test_analyze_bill_file_single_failure_writes_unknown(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    from beartools.bill.service import analyze_bill_file

    input_path = tmp_path / "test.xlsx"
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["原始来源", "交易时间", "交易对方", "金额", "交易状态", "注意", "备注"])
    ws.append(["2601-支付宝", "2025-12-31 18:54:05", "山姆会员商店", "289.80", "交易成功", "", "日用百货"])
    wb.save(input_path)

    def mock_row_analyzer(*args, **kwargs) -> None:
        raise RuntimeError("analyze failed")

    result = analyze_bill_file(input_path, row_analyzer=mock_row_analyzer)
    assert result.total_rows == 1
    assert result.failed_rows == 1
    assert result.output_path.exists()

    wb_out = load_workbook(result.output_path)
    ws_out = wb_out.active
    row2 = [cell.value for cell in ws_out[2]]
    assert row2[-2] == "unknow"
    assert row2[-1] == "unknow"


def test_analyze_bill_file_5_failures_still_success(tmp_path: Path) -> None:
    from beartools.bill.service import analyze_bill_file

    input_path = tmp_path / "test.xlsx"
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["原始来源", "交易时间", "交易对方", "金额", "交易状态", "注意", "备注"])
    for i in range(5):
        ws.append(["2601-支付宝", f"2025-12-3{i} 18:54:05", f"商家{i}", f"{10 * (i + 1)}.00", "交易成功", "", ""])
    wb.save(input_path)

    def mock_row_analyzer(*args, **kwargs) -> None:
        raise RuntimeError("analyze failed")

    result = analyze_bill_file(input_path, row_analyzer=mock_row_analyzer)
    assert result.total_rows == 5
    assert result.failed_rows == 5
    assert result.output_path.exists()


def test_analyze_bill_file_6_failures_raises_and_no_output(tmp_path: Path) -> None:
    from beartools.bill.service import analyze_bill_file

    input_path = tmp_path / "test.xlsx"
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["原始来源", "交易时间", "交易对方", "金额", "交易状态", "注意", "备注"])
    for i in range(6):
        ws.append(["2601-支付宝", f"2025-12-3{i} 18:54:05", f"商家{i}", f"{10 * (i + 1)}.00", "交易成功", "", ""])
    wb.save(input_path)

    def mock_row_analyzer(*args, **kwargs) -> None:
        raise RuntimeError("analyze failed")

    import pytest

    output_path = input_path.with_name(f"{input_path.stem}.analysis.xlsx")
    assert not output_path.exists()

    with pytest.raises(RuntimeError, match="失败行数超过阈值"):
        analyze_bill_file(input_path, row_analyzer=mock_row_analyzer)

    assert not output_path.exists(), "出错时不应留下输出文件"


def test_analyze_bill_file_non_xlsx_raises(tmp_path: Path) -> None:
    import pytest

    from beartools.bill.service import analyze_bill_file

    input_path = tmp_path / "test.csv"
    input_path.write_text("test", encoding="utf-8")

    with pytest.raises(ValueError, match="只支持归一化结果 .xlsx 输入"):
        analyze_bill_file(input_path)


def test_analyze_bill_file_missing_required_headers_raises(tmp_path: Path) -> None:
    import pytest

    from beartools.bill.service import analyze_bill_file

    input_path = tmp_path / "test.xlsx"
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["原始来源", "交易时间", "交易对方", "金额"])  # 缺少几个表头
    ws.append(["2601-支付宝", "2025-12-31 18:54:05", "山姆会员商店", "289.80"])
    wb.save(input_path)

    with pytest.raises(RuntimeError, match="表头不匹配，需要包含固定表头"):
        analyze_bill_file(input_path)


def test_run_bill_pipeline_success(tmp_path: Path):
    from beartools.bill.models import (
        BillAnalysisResult,
        BillFieldDetail,
        BillFieldMapping,
        BillRemarkColumns,
        BillStructureFileResult,
    )
    from beartools.bill.service import run_bill_pipeline

    input_csv = tmp_path / "test.csv"
    input_csv.write_text(
        "导出信息\n交易时间,交易分类,交易对方,金额,交易状态,收/支,备注\n"
        "2025-12-31 18:54:05,日用百货,山姆会员商店,289.80,交易成功,支出,测试备注\n",
        encoding="utf-8",
    )
    structure = BillStructureFileResult(
        file_name="test.csv",
        source="测试账单",
        header_row=2,
        data_start_row=3,
        field_mapping=BillFieldMapping(
            transaction_time=BillFieldDetail(column_name="交易时间", confidence="high", reason=""),
            counterparty=BillFieldDetail(column_name="交易对方", confidence="high", reason=""),
            amount=BillFieldDetail(column_name="金额", confidence="high", reason=""),
            status=BillFieldDetail(column_name="交易状态", confidence="high", reason=""),
            remark_columns=BillRemarkColumns(column_names=["交易分类", "收/支", "备注"], confidence="high", reason=""),
        ),
    )
    cwd = Path.cwd()
    os.chdir(tmp_path)
    try:
        result = run_bill_pipeline(
            input_csv,
            "2601-",
            structure_resolver=lambda _: structure,
            row_analyzer=lambda *args: BillAnalysisResult(purpose="测试用途", owner="测试归属"),
        )
        assert result.normalized_output_path.exists()
        assert result.analysis_output_path.exists()
        assert result.normalized_row_count == 1
        assert result.analysis_total_rows == 1
        assert result.analysis_failed_rows == 0
    finally:
        os.chdir(cwd)


def test_run_bill_pipeline_analysis_fails(tmp_path: Path):
    from beartools.bill.models import BillFieldDetail, BillFieldMapping, BillRemarkColumns, BillStructureFileResult
    from beartools.bill.service import run_bill_pipeline

    input_csv = tmp_path / "test.csv"
    input_csv.write_text(
        "导出信息\n交易时间,交易分类,交易对方,金额,交易状态,收/支,备注\n"
        "2025-12-31 18:54:05,日用百货,山姆会员商店,289.80,交易成功,支出,测试备注\n"
        "2025-12-31 18:55:05,日用百货,山姆会员商店,289.80,交易成功,支出,测试备注\n"
        "2025-12-31 18:56:05,日用百货,山姆会员商店,289.80,交易成功,支出,测试备注\n"
        "2025-12-31 18:57:05,日用百货,山姆会员商店,289.80,交易成功,支出,测试备注\n"
        "2025-12-31 18:58:05,日用百货,山姆会员商店,289.80,交易成功,支出,测试备注\n"
        "2025-12-31 18:59:05,日用百货,山姆会员商店,289.80,交易成功,支出,测试备注\n"
        "2025-12-31 19:00:05,日用百货,山姆会员商店,289.80,交易成功,支出,测试备注\n",
        encoding="utf-8",
    )
    structure = BillStructureFileResult(
        file_name="test.csv",
        source="测试账单",
        header_row=2,
        data_start_row=3,
        field_mapping=BillFieldMapping(
            transaction_time=BillFieldDetail(column_name="交易时间", confidence="high", reason=""),
            counterparty=BillFieldDetail(column_name="交易对方", confidence="high", reason=""),
            amount=BillFieldDetail(column_name="金额", confidence="high", reason=""),
            status=BillFieldDetail(column_name="交易状态", confidence="high", reason=""),
            remark_columns=BillRemarkColumns(column_names=["交易分类", "收/支", "备注"], confidence="high", reason=""),
        ),
    )
    cwd = Path.cwd()
    os.chdir(tmp_path)
    try:

        def failing_row_analyzer(*args):
            raise RuntimeError("分析失败")

        run_bill_pipeline(input_csv, "2601-", structure_resolver=lambda _: structure, row_analyzer=failing_row_analyzer)
        raise AssertionError("应该抛出异常")
    except RuntimeError:
        normalized_path = Path("data/bill/2601-测试账单.xlsx")
        analysis_path = Path("data/bill/2601-测试账单.analysis.xlsx")
        assert normalized_path.exists(), "normalized应该保留"
        assert not analysis_path.exists(), "analysis不应该存在"
    finally:
        os.chdir(cwd)


def test_run_bill_pipeline_updates_progress_state(tmp_path: Path) -> None:
    import os

    from beartools.bill.models import (
        BillAnalysisResult,
        BillFieldDetail,
        BillFieldMapping,
        BillRemarkColumns,
        BillRunProgressState,
        BillStructureFileResult,
    )
    from beartools.bill.service import run_bill_pipeline

    input_csv = tmp_path / "test.csv"
    input_csv.write_text(
        "导出信息\n交易时间,交易分类,交易对方,金额,交易状态,收/支,备注\n"
        "2025-12-31 18:54:05,日用百货,山姆会员商店,289.80,交易成功,支出,测试备注\n",
        encoding="utf-8",
    )
    structure = BillStructureFileResult(
        file_name="test.csv",
        source="测试账单",
        header_row=2,
        data_start_row=3,
        field_mapping=BillFieldMapping(
            transaction_time=BillFieldDetail(column_name="交易时间", confidence="high", reason=""),
            counterparty=BillFieldDetail(column_name="交易对方", confidence="high", reason=""),
            amount=BillFieldDetail(column_name="金额", confidence="high", reason=""),
            status=BillFieldDetail(column_name="交易状态", confidence="high", reason=""),
            remark_columns=BillRemarkColumns(column_names=["交易分类", "收/支", "备注"], confidence="high", reason=""),
        ),
    )
    progress_state = BillRunProgressState()

    cwd = Path.cwd()
    os.chdir(tmp_path)
    try:
        result = run_bill_pipeline(
            input_csv,
            "2601-",
            structure_resolver=lambda _: structure,
            row_analyzer=lambda *args: BillAnalysisResult(purpose="测试用途", owner="vv"),
            progress_state=progress_state,
        )
    finally:
        os.chdir(cwd)

    assert result.analysis_total_rows == 1
    assert progress_state.current_step == "Finished"
    assert progress_state.analysis_completed_count == 1


def test_analyze_bill_file_counts_completed_rows_even_when_row_falls_back_to_unknow(tmp_path: Path) -> None:
    from openpyxl import Workbook

    from beartools.bill.models import BillRunProgressState
    from beartools.bill.service import analyze_bill_file

    input_path = tmp_path / "test.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.append(["原始来源", "交易时间", "交易对方", "金额", "交易状态", "注意", "备注"])
    worksheet.append(["2601-支付宝", "2025-12-31 18:54:05", "山姆会员商店", "289.80", "交易成功", "", "日用百货"])
    worksheet.append(["2601-支付宝", "2025-12-31 18:55:05", "盒马", "19.80", "交易成功", "", "零食"])
    workbook.save(input_path)

    progress_state = BillRunProgressState()

    def mixed_row_analyzer(counterparty: str, remark: str, status: str, amount: str):
        if counterparty == "盒马":
            raise RuntimeError("分析失败")
        from beartools.bill.models import BillAnalysisResult

        return BillAnalysisResult(purpose="食物", owner="yy")

    result = analyze_bill_file(input_path, row_analyzer=mixed_row_analyzer, progress_state=progress_state)

    assert result.total_rows == 2
    assert result.failed_rows == 1
    assert progress_state.current_step == "Analysis"
    assert progress_state.analysis_total_count == 2
    assert progress_state.analysis_completed_count == 2
